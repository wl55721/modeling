"""Component classification and color palette for Excel output."""
from __future__ import annotations

from typing import Optional

from openpyxl.styles import PatternFill

FILL_COLORS = {
    "attn_norm":    "E8F5E9",
    "ffn_norm":     "E8F5E9",
    "final_norm":   "E8F5E9",
    "attn.":        "E3F2FD",
    "moe.gate":     "FFF3E0",
    "moe.shared":   "FFF8E1",
    "moe.expert":   "FCE4EC",
    "ffn":          "F3E5F5",
    "embedding":    "ECEFF1",
    "lm_head":      "ECEFF1",
    "add":          "FFFFFF",
}


def extract_layer_idx(module_path: str) -> str:
    parts = module_path.split(".")
    for i, p in enumerate(parts):
        if p == "layers" and i + 1 < len(parts):
            try:
                return str(int(parts[i + 1]))
            except ValueError:
                pass
    return ""


def classify_component(module_path: str, func_name: str) -> str:
    """Map a module path to a human-readable component category."""
    s = module_path.lower()
    fn_parts = func_name.split(".")
    op_short = fn_parts[1] if len(fn_parts) >= 2 else func_name

    if "input_layernorm" in s:
        return "attn_norm"
    if "post_attention_layernorm" in s:
        return "ffn_norm"

    if "q_a_proj" in s:
        return "attn.q_a_proj"
    if "q_a_layernorm" in s:
        return "attn.q_norm"
    if "q_b_proj" in s:
        return "attn.q_b_proj"
    if "q_proj" in s:
        return "attn.q_proj"
    if "kv_a_proj" in s:
        return "attn.kv_a_proj"
    if "kv_a_layernorm" in s:
        return "attn.kv_norm"
    if "kv_b_proj" in s:
        return "attn.kv_b_proj"
    if "o_proj" in s:
        return "attn.o_proj"

    if "self_attn" in s or "attn" in s:
        if "rotary" in s:
            return "attn.rope"
        if op_short in ("matmul", "mm", "bmm"):
            return "attn.score"
        if "softmax" in op_short or "safe_softmax" in op_short:
            return "attn.softmax"
        return f"attn.{op_short}"

    if "gate" in s and "experts" not in s and "up" not in s and ("moe" in s or "mlp" in s):
        return f"moe.gate.{op_short}"
    if "shared_expert" in s or ("shared" in s and "mlp" in s):
        if "gate_proj" in s:
            return "moe.shared.gate_proj"
        if "up_proj" in s:
            return "moe.shared.up_proj"
        if "down_proj" in s:
            return "moe.shared.down_proj"
        return f"moe.shared.{op_short}"
    if "experts" in s:
        return f"moe.experts.{op_short}"

    if "mlp" in s or "moe" in s:
        if "gate_proj" in s:
            return "ffn.gate_proj"
        if "up_proj" in s:
            return "ffn.up_proj"
        if "down_proj" in s:
            return "ffn.down_proj"
        if op_short == "silu":
            return "ffn.silu"
        if op_short == "mul":
            return "ffn.mul"
        return f"ffn.{op_short}"

    if "embed" in s:
        return "embedding"
    if "norm" in s:
        return "final_norm"
    if "lm_head" in s:
        return "lm_head"

    return op_short


def get_fill(component: str) -> Optional[PatternFill]:
    for prefix, color in FILL_COLORS.items():
        if component.startswith(prefix):
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
    return None
