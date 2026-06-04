"""Multi-configuration comparison reports."""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from python.zrt.report.summary import E2ESummary

logger = logging.getLogger(__name__)


@dataclass
class ComparisonReport:
    """Comparison of multiple configurations for the same model + hardware + phase."""
    model: str
    hardware: str
    phase: str
    entries: list[tuple[str, "E2ESummary"]] = field(default_factory=list)

    @property
    def config_labels(self) -> list[str]:
        return [label for label, _ in self.entries]

    def metric_table(self, metrics: list[str] | None = None) -> dict[str, list]:
        """Return {metric: [values per config]} for key metrics."""
        if metrics is None:
            metrics = [
                "latency_ms", "tokens_per_sec", "mfu",
                "compute_ms", "comm_ms", "overlap_ratio",
                "total_flops", "arithmetic_intensity",
            ]
        table: dict[str, list] = {m: [] for m in metrics}
        for _, summary in self.entries:
            for m in metrics:
                val = getattr(summary, m, None)
                if isinstance(val, float):
                    table[m].append(round(val, 4))
                else:
                    table[m].append(val)
        return table


def build_comparison_report(
    entries: list[tuple[str, "E2ESummary"]],
) -> ComparisonReport:
    """Build a comparison report from multiple (label, E2ESummary) pairs.

    Parameters
    ----------
    entries : list[tuple[str, E2ESummary]]
        Each pair is (config_description, summary), e.g.
        ``[("TP=1", summary_tp1), ("TP=4", summary_tp4)]``.

    Returns
    -------
    ComparisonReport
    """
    if not entries:
        raise ValueError("At least one entry is required")
    first = entries[0][1]
    return ComparisonReport(
        model=first.model,
        hardware=first.hardware,
        phase=first.phase,
        entries=entries,
    )


def export_comparison_excel(
    report: ComparisonReport,
    output_path: Path,
) -> Path:
    """Export comparison report to Excel with overview and scaling sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin", color="BDBDBD"))

    # ── Sheet 1: Comparison Overview ──
    ws = wb.active
    ws.title = "Comparison Overview"
    ws.column_dimensions["A"].width = 24

    labels = report.config_labels
    for i, label in enumerate(labels, 2):
        ws.cell(row=1, column=i, value=label).font = header_font
        ws.cell(row=1, column=i).fill = header_fill
        ws.cell(row=1, column=i).alignment = Alignment(horizontal="center")
    ws.column_dimensions["B"].width = 16

    metrics = [
        ("Latency (ms)", "latency_ms", 3),
        ("Throughput (tok/s)", "tokens_per_sec", 1),
        ("TTFT (ms)", "ttft_ms", 3),
        ("TPOT (ms/token)", "tpot_ms", 3),
        ("MFU", "mfu", 4),
        ("HBM BW util", "hbm_bandwidth_util", 4),
        ("Compute (ms)", "compute_ms", 3),
        ("Comm (ms)", "comm_ms", 3),
        ("Exposed comm (ms)", "exposed_comm_ms", 3),
        ("Overlap ratio", "overlap_ratio", 4),
        ("Total FLOPs (T)", "total_flops", 3, lambda v: v / 1e12),
        ("Total bytes (GB)", "total_bytes", 3, lambda v: v / 1e9),
        ("Arith intensity", "arithmetic_intensity", 2),
    ]

    ws.cell(row=1, column=1, value="Metric").font = header_font
    ws.cell(row=1, column=1).fill = header_fill

    for row_idx, (label, attr, decimals, *scale_fn) in enumerate(metrics, 2):
        ws.cell(row=row_idx, column=1, value=label).border = thin_border
        for col_idx, (_, summary) in enumerate(report.entries, 2):
            val = getattr(summary, attr, None)
            if val is not None:
                if scale_fn:
                    val = scale_fn[0](val)
                if isinstance(val, float) and attr in ("mfu", "hbm_bandwidth_util", "overlap_ratio"):
                    val = f"{val:.2%}"
                elif isinstance(val, float):
                    val = round(val, decimals)
            ws.cell(row=row_idx, column=col_idx, value=val).border = thin_border

    # ── Sheet 2: Scaling Efficiency ──
    ws2 = wb.create_sheet("Scaling Efficiency")
    ws2.column_dimensions["A"].width = 24
    ws2.cell(row=1, column=1, value="Metric").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    for i, label in enumerate(labels, 2):
        ws2.cell(row=1, column=i, value=label).font = header_font
        ws2.cell(row=1, column=i).fill = header_fill
        ws2.cell(row=1, column=i).alignment = Alignment(horizontal="center")

    baseline_lat = report.entries[0][1].latency_ms if report.entries else 1
    scaling_metrics = [
        ("Speedup", "latency_ms", True),
        ("Efficiency", "latency_ms", True),
        ("MFU", "mfu", False),
        ("Throughput (tok/s)", "tokens_per_sec", False),
    ]

    for row_idx, (label, attr, invert) in enumerate(scaling_metrics, 2):
        ws2.cell(row=row_idx, column=1, value=label).border = thin_border
        for col_idx, (_, summary) in enumerate(report.entries, 2):
            val = getattr(summary, attr, None)
            if val is not None and baseline_lat > 0:
                if invert:
                    ratio = baseline_lat / val
                else:
                    ratio = val
                if label == "Efficiency":
                    # Efficiency = speedup / relative_configs
                    ratio = (baseline_lat / val) / col_idx if col_idx > 0 else 0
                ws2.cell(row=row_idx, column=col_idx,
                         value=round(ratio, 4)).border = thin_border
            else:
                ws2.cell(row=row_idx, column=col_idx, value="").border = thin_border

    ws2.freeze_panes = "A2"
    wb.save(output_path)
    logger.info("Exported comparison Excel to %s", output_path)
    return output_path


def export_comparison_html(
    report: ComparisonReport,
    output_path: Path,
) -> Path:
    """Export comparison report to interactive HTML with bar charts."""
    labels = report.config_labels
    table = report.metric_table()

    # Build bar chart data for latency
    latencies = table.get("latency_ms", [])
    max_lat = max(latencies) if latencies else 1

    bars_html = ""
    for i, (label, lat) in enumerate(zip(labels, latencies)):
        pct = (lat / max_lat * 100) if max_lat > 0 else 0
        color = "#1a237e" if i == 0 else "#3949ab"
        bars_html += (
            f'<div style="margin:8px 0">'
            f'<span style="display:inline-block;width:80px;font-size:13px">{label}</span>'
            f'<div style="display:inline-block;width:300px;height:24px;background:#e0e0e0;border-radius:4px;overflow:hidden">'
            f'<div style="width:{pct}%;height:100%;background:{color};display:flex;align-items:center;padding-left:8px;'
            f'color:#fff;font-size:12px;font-weight:600">{lat:.2f}ms</div></div></div>'
        )

    # MFU comparison
    mfus = table.get("mfu", [])
    mfu_bars = ""
    max_mfu = max(mfus) if mfus else 1
    for i, (label, mfu) in enumerate(zip(labels, mfus)):
        pct = (mfu / max_mfu * 100) if max_mfu > 0 else 0
        color = "#4CAF50" if i == 0 else "#66BB6A"
        mfu_bars += (
            f'<div style="margin:8px 0">'
            f'<span style="display:inline-block;width:80px;font-size:13px">{label}</span>'
            f'<div style="display:inline-block;width:300px;height:24px;background:#e0e0e0;border-radius:4px;overflow:hidden">'
            f'<div style="width:{pct}%;height:100%;background:{color};display:flex;align-items:center;padding-left:8px;'
            f'color:#fff;font-size:12px;font-weight:600">{mfu:.1%}</div></div></div>'
        )

    # Throughput comparison
    throughputs = table.get("tokens_per_sec", [])
    tp_bars = ""
    max_tp = max(throughputs) if throughputs else 1
    for i, (label, tp) in enumerate(zip(labels, throughputs)):
        pct = (tp / max_tp * 100) if max_tp > 0 else 0
        color = "#FF9800" if i == 0 else "#FFB74D"
        tp_bars += (
            f'<div style="margin:8px 0">'
            f'<span style="display:inline-block;width:80px;font-size:13px">{label}</span>'
            f'<div style="display:inline-block;width:300px;height:24px;background:#e0e0e0;border-radius:4px;overflow:hidden">'
            f'<div style="width:{pct}%;height:100%;background:{color};display:flex;align-items:center;padding-left:8px;'
            f'color:#fff;font-size:12px;font-weight:600">{tp:.0f}</div></div></div>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Comparison: {report.model} | {report.phase.upper()}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
       background: #f5f5f5; color: #333; padding: 24px; }}
h1 {{ font-size: 22px; margin-bottom: 8px; color: #1a237e; }}
h2 {{ font-size: 16px; margin: 24px 0 12px; color: #37474f; border-bottom: 2px solid #1a237e; padding-bottom: 4px; }}
.meta {{ font-size: 13px; color: #78909c; margin-bottom: 16px; }}
.card {{ background: #fff; border-radius: 8px; padding: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.12); margin-bottom: 16px; }}
table {{ width: 100%; border-collapse: collapse; background: #fff; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.12); }}
th {{ background: #1a237e; color: #fff; padding: 10px 12px; text-align: left; font-size: 13px; }}
td {{ padding: 8px 12px; border-bottom: 1px solid #e0e0e0; font-size: 13px; }}
tr:hover td {{ background: #f5f5f5; }}
</style>
</head>
<body>
<h1>Comparison: {report.model} | {report.phase.upper()}</h1>
<div class="meta">Hardware: {report.hardware} | Configurations: {', '.join(labels)}</div>

<h2>Latency Comparison</h2>
<div class="card">{bars_html}</div>

<h2>MFU Comparison</h2>
<div class="card">{mfu_bars}</div>

<h2>Throughput Comparison (tokens/s)</h2>
<div class="card">{tp_bars}</div>

<h2>Detailed Metrics</h2>
<table>
<thead><tr><th>Metric</th>{''.join(f'<th>{l}</th>' for l in labels)}</tr></thead>
<tbody>
<tr><td>Latency (ms)</td>{''.join(f'<td>{v:.3f}</td>' for v in latencies)}</tr>
<tr><td>Throughput (tok/s)</td>{''.join(f'<td>{v:.0f}</td>' for v in throughputs)}</tr>
<tr><td>MFU</td>{''.join(f'<td>{v:.2%}</td>' for v in mfus)}</tr>
<tr><td>Compute (ms)</td>{''.join(f'<td>{v:.3f}</td>' for v in table.get("compute_ms", []))}</tr>
<tr><td>Comm (ms)</td>{''.join(f'<td>{v:.3f}</td>' for v in table.get("comm_ms", []))}</tr>
<tr><td>Overlap ratio</td>{''.join(f'<td>{v:.1%}</td>' for v in table.get("overlap_ratio", []))}</tr>
</tbody>
</table>
</body>
</html>"""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    logger.info("Exported comparison HTML to %s", output_path)
    return output_path
