"""Integration test: verify how DataParallel (DP) affects TrainingReport fields.

This test runs the training modelling CLI for dp=1, dp=4, dp=8, loads the
generated ``deepseek_v4_training_report.json`` files and asserts expected
relationships:

- optimizer state (opt_state) per-GPU ≈ 1/dp  (ZeRO-1)
- total per-GPU memory drops monotonically as dp increases
- step_time decreases and tokens/sec increases monotonically as dp increases
- dp_hidden_ms and dp_exposed_ms are zero for dp=1, non-zero for dp>1
- dp_comm_total (dp_hidden + dp_exposed) increases monotonically with dp
- optimizer state scales approximately as 1/dp across dp=4 and dp=8
- DP communication hiding invariants (dp_total = dp_hidden + dp_exposed, etc.)

This is a long-running integration test that captures real reports. To avoid
running it by default in fast CI, it is skipped unless the environment
variable ``RUN_DP_TEST`` is set to ``1``.

Run locally (PowerShell):

```powershell
$env:RUN_DP_TEST='1'; pytest tests/IT/test_dp_effects.py -q
```
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pytest

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class _DPResult:
    """Report dict + output directory, with dict-like access for backward compat."""

    def __init__(self, report: dict, out_dir: Path):
        self._report = report
        self.out_dir = out_dir

    def __getitem__(self, key):
        return self._report[key]

    def get(self, key, default=None):
        return self._report.get(key, default)


def _build_training_cli_cmd(
    outdir: Path,
    *,
    dp: int,
    global_batch: int = 64,
    dp_ddp_buckets: bool = True,
    dp_bucket_cap_mb: float = 25,
) -> list[str]:
    cmd = [
        sys.executable,
        "-m",
        "python.zrt",
        "--model-id", "hf_models/deepseek_v4",
        "--train",
        "--hw", "nvidia_h100_sxm",
        "--hidden", "7168",
        "--layers", "4",
        "--seq-len", "128",
        "--global-batch", str(global_batch),
        "--micro-batch", "8",
        "--dp", str(dp),
        "--pp", "4",
        "--tp", "1",
        "--pp-schedule", "1f1b",
        "--recompute-policy", "full",
        "--optimizer", "adam",
        "--zero-stage", "1",
    ]
    if dp_ddp_buckets:
        cmd.extend(["--dp-ddp-buckets", "--dp-bucket-cap-mb", f"{dp_bucket_cap_mb:g}"])
    cmd.extend(["--output-dir", str(outdir)])
    return cmd


def _run_cli_and_load_report(
    repo_root: Path,
    outdir: Path,
    dp: int,
    timeout: int = 900,
    *,
    global_batch: int = 64,
) -> _DPResult:
    """Run `python -m python.zrt` with given dp and return report + out_dir.

    Raises subprocess.CalledProcessError on failure.
    """
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "python")

    cmd = _build_training_cli_cmd(outdir, dp=dp, global_batch=global_batch)
    proc = subprocess.run(cmd, check=True, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, timeout=timeout)
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "cli_output.log").write_text(proc.stdout)

    report_path = outdir / "reports" / "deepseek_v4_training_report.json"
    assert report_path.exists(), f"Report not found at {report_path}"
    return _DPResult(json.loads(report_path.read_text()), outdir)


@pytest.fixture(scope="session")
def dp_reports(tmp_path_factory):
    """Session-scoped fixture: run CLI for dp=1, dp=4, dp=8, return all reports."""
    if os.environ.get("RUN_DP_TEST") != "1":
        pytest.skip("Set RUN_DP_TEST=1 to run this long integration test")

    repo_root = Path(__file__).resolve().parents[2]
    tmp_path = tmp_path_factory.mktemp("dp_effects")

    reports = {}
    for dp in (1, 4, 8):
        out_dir = tmp_path / f"out_dp{dp}"
        reports[dp] = _run_cli_and_load_report(repo_root, out_dir, dp=dp)

    return reports


def test_requested_dp_ddp_bucket_command_e2e():
    """Smoke the exact DDP-bucket CLI shape requested for pp=4/dp=4.

    This is skipped with the rest of the long DP integration tests unless
    RUN_DP_TEST=1 is set.
    """
    if os.environ.get("RUN_DP_TEST") != "1":
        pytest.skip("Set RUN_DP_TEST=1 to run this long integration test")

    repo_root = Path(__file__).resolve().parents[2]
    out_dir = repo_root / "output" / "dp_test_pp4_dp4"
    result = _run_cli_and_load_report(
        repo_root,
        out_dir,
        dp=4,
        global_batch=32,
    )

    report = result._report
    assert report.get("dp_total_ms", 0.0) > 0
    assert (out_dir / "pp_trace" / "pp_per_stage.json").exists()
    assert (out_dir / "pp_trace" / "pp_combined.json").exists()
    assert (out_dir / "pp_trace" / "pp_stitched.json").exists()


def _load_xlsx_sheets(xlsx_path: Path) -> dict[str, list[dict]]:
    """Load all sheets from training xlsx, return {sheet_name: [row_dict]}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        headers = [
            str(h) if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        data = []
        for row in rows[1:]:
            vals = [v for v in row]
            if all(v is None or v == "" for v in vals):
                continue
            data.append(dict(zip(headers, vals)))
        result[sheet_name] = data
    wb.close()
    return result


def _filter_dp_comm(comm_sheet: list[dict]) -> list[dict]:
    """Return only DP gradient reduction communication rows."""
    return [
        r for r in comm_sheet
        if r.get("Role") == "dp_grad_reduce"
    ]


@pytest.fixture(scope="session")
def dp_excel_data(dp_reports):
    """Session fixture: load Excel sheets from CLI runs cached by dp_reports."""
    if not HAS_OPENPYXL:
        pytest.skip("openpyxl required: pip install openpyxl")

    data = {}
    for dp in (1, 4):
        out_dir = dp_reports[dp].out_dir
        xlsx_path = out_dir / "deepseek_v4_training.xlsx"
        assert xlsx_path.exists(), f"Excel not found at {xlsx_path}"
        data[dp] = _load_xlsx_sheets(xlsx_path)

    return data


# ── ZeRO-1 optimizer state scaling ─────────────────────────────────────────

def test_dp_optimizer_state_scales_inverse(dp_reports):
    """Optimizer state per-GPU should roughly scale ~1/dp (ZeRO-1 behaviour)."""
    opt = {}
    for dp in (1, 4, 8):
        mb = dp_reports[dp].get("memory_breakdown_gb")
        assert mb, f"memory_breakdown_gb missing for dp={dp}"
        opt[dp] = mb.get("opt_state")
        assert opt[dp] is not None, f"opt_state missing for dp={dp}"

    # Approximate 1/dp scaling: opt(dp) ≈ opt(1) / dp
    for dp in (4, 8):
        expected = opt[1] / dp
        assert opt[dp] == pytest.approx(expected, rel=0.25), (
            f"opt_state did not scale near 1/dp: dp=1 → {opt[1]}, dp={dp} → {opt[dp]}, "
            f"expected ≈ {expected}"
        )

    # Monotonic decrease
    assert opt[1] > opt[4] > opt[8], (
        f"opt_state should decrease monotonically: dp1={opt[1]}, dp4={opt[4]}, dp8={opt[8]}"
    )


# ── Total memory monotonicity ──────────────────────────────────────────────

def test_dp_total_memory_decreases(dp_reports):
    """Total per-GPU memory should decrease monotonically as dp increases."""
    total = {}
    for dp in (1, 4, 8):
        mb = dp_reports[dp]["memory_breakdown_gb"]
        total[dp] = mb["total"]
        assert total[dp] is not None

    assert total[1] > total[4] > total[8], (
        f"total memory should decrease monotonically: "
        f"dp1={total[1]}, dp4={total[4]}, dp8={total[8]}"
    )


# ── Throughput monotonicity ────────────────────────────────────────────────

def test_dp_throughput_improves_over_dp1(dp_reports):
    """Multi-DP should improve over dp=1, but dp8 need not beat dp4 here.

    The fixture keeps global_batch fixed at 64 with micro_batch=8 and pp=4.
    That gives M=2 for dp=4 and M=1 for dp=8; the latter can have a larger
    pipeline bubble, so strict dp1 > dp4 > dp8 monotonicity is not a valid
    invariant for this geometry.
    """
    step_time = {dp: dp_reports[dp]["step_time_ms"] for dp in (1, 4, 8)}

    assert step_time[1] > step_time[4], (
        f"dp=4 step_time should improve over dp=1: "
        f"dp1={step_time[1]}, dp4={step_time[4]}, dp8={step_time[8]}"
    )
    assert step_time[1] > step_time[8], (
        f"dp=8 step_time should improve over dp=1 even if PP bubble makes it "
        f"slower than dp=4: dp1={step_time[1]}, dp4={step_time[4]}, dp8={step_time[8]}"
    )


# ── DP communication accounting ────────────────────────────────────────────

def test_dp_communication_zero_for_dp1(dp_reports):
    """dp_hidden_ms and dp_exposed_ms should be 0 for dp=1 (no DP communication)."""
    rep1 = dp_reports[1]

    dp_hidden1 = rep1.get("dp_hidden_ms")
    dp_exposed1 = rep1.get("dp_exposed_ms")
    assert dp_hidden1 is not None
    assert dp_exposed1 is not None
    assert dp_hidden1 == 0.0, f"dp_hidden should be 0 for dp=1, got {dp_hidden1}"
    assert dp_exposed1 == 0.0, f"dp_exposed should be 0 for dp=1, got {dp_exposed1}"


def test_dp_communication_nonzero_for_dp_gt1(dp_reports):
    """dp_hidden + dp_exposed should be >0 for dp>1 (DP AR/RS communication exists).

    dp_hidden may be 0 (e.g. no bubble to absorb AR) or dp_exposed may be 0
    (e.g. AR fully hidden in bubble), but their sum must reflect the total
    DP communication volume.
    """
    for dp in (4, 8):
        rep = dp_reports[dp]
        dp_hidden = rep.get("dp_hidden_ms")
        dp_exposed = rep.get("dp_exposed_ms")
        dp_comm = rep.get("dp_total_ms", 0.0)
        assert dp_hidden is not None, f"dp_hidden_ms missing for dp={dp}"
        assert dp_exposed is not None, f"dp_exposed_ms missing for dp={dp}"
        assert dp_comm > 0, (
            f"dp_hidden + dp_exposed should be >0 for dp={dp}, "
            f"got hidden={dp_hidden}, exposed={dp_exposed}"
        )


def test_dp_comm_volume_monotonic(dp_reports):
    """Total DP communication volume (dp_hidden + dp_exposed) should increase with dp.

    Larger DP group means more gradient data to reduce, so the total DP
    communication time (exposed + hidden) should grow monotonically.
    Note: on full-mesh topologies the per-step time may not scale linearly
    due to the (N-1)/N ring factor, but the trend should be increasing.
    """
    dp_comm = {}
    for dp in (4, 8):
        rep = dp_reports[dp]
        dp_comm[dp] = rep.get("dp_total_ms", 0.0)

    assert dp_comm[8] > dp_comm[4], (
        f"DP comm volume should increase with dp: "
        f"dp4={dp_comm[4]:.2f}ms, dp8={dp_comm[8]:.2f}ms"
    )


# ── DP communication hiding ────────────────────────────────────────────────

class TestDpCommHiding:
    """DP gradient reduction hiding in pipeline bubble.

    Core formula (schedules.py _dp_hidden / _dp_hide_window):

        hide_window = cooldown + ratio * steady_bwd_total   (if dp_overlap_in_bubble)
        max_hidable = dp_ar_time * (1 - 1/dp_grad_buckets)  (last bucket always exposed)
        dp_hidden   = min(hide_window, max_hidable)
        dp_exposed  = dp_ar_time - dp_hidden

    After augmentation, dp_hidden is recalculated as:
        dp_hidden = max(0, dp_ar_time - dp_exposed)

    Key invariants (Stack B / trace mode):
      1. dp_total_ms = dp_hidden_ms + dp_exposed_ms
      2. dp_hidden_ms <= bubble_time_ms (cannot hide more than bubble allows)
      3. dp_exposed_ms >= dp_total_ms / dp_grad_buckets (last bucket residual)
      4. step_time = pipeline_time + dp_exposed + optimizer_time + optimizer_comm
      5. pipeline_time = warmup + steady + cooldown  (dp_exposed NOT included)
      6. dp_hidden is absorbed in bubble, does NOT add to step_time
    """

    def test_dp_total_equals_hidden_plus_exposed(self, dp_reports):
        """dp_total_ms must equal dp_hidden_ms + dp_exposed_ms for all dp."""
        for dp in (1, 4, 8):
            rep = dp_reports[dp]
            total = rep.get("dp_total_ms", 0.0)
            hidden = rep.get("dp_hidden_ms", 0.0)
            exposed = rep.get("dp_exposed_ms", 0.0)
            assert total == pytest.approx(hidden + exposed, abs=1e-9), (
                f"dp={dp}: dp_total({total}) != dp_hidden({hidden}) + dp_exposed({exposed})"
            )

    def test_dp_hidden_leq_bubble_time(self, dp_reports):
        """dp_hidden cannot exceed the pipeline bubble time (warmup + cooldown)."""
        for dp in (4, 8):
            rep = dp_reports[dp]
            dp_hidden = rep.get("dp_hidden_ms", 0.0)
            bubble = rep.get("bubble_time_ms", 0.0)
            assert dp_hidden <= bubble + 1e-6, (
                f"dp={dp}: dp_hidden({dp_hidden:.4f}ms) > bubble_time({bubble:.4f}ms)"
            )

    def test_dp_exposed_ge_last_bucket_residual(self, dp_reports):
        """The last gradient bucket's collective is always exposed.

        dp_exposed >= dp_total / dp_grad_buckets (default buckets=25).
        """
        dp_grad_buckets = 25
        for dp in (4, 8):
            rep = dp_reports[dp]
            total = rep.get("dp_total_ms", 0.0)
            exposed = rep.get("dp_exposed_ms", 0.0)
            min_exposed = total / dp_grad_buckets
            assert exposed >= min_exposed - 1e-9, (
                f"dp={dp}: dp_exposed({exposed:.6f}ms) < "
                f"dp_total/buckets({min_exposed:.6f}ms)"
            )

    def test_step_time_equals_pipeline_plus_dp_exposed_plus_optimizer(self, dp_reports):
        """step_time = pipeline_time + dp_exposed + optimizer_time + optimizer_comm.

        Stack B (trace mode): pipeline_time = warmup + steady + cooldown
        (does NOT include dp_exposed). dp_exposed is added to step_time
        separately, then optimizer time is appended.
        """
        for dp in (1, 4, 8):
            rep = dp_reports[dp]
            step = rep["step_time_ms"]
            pipeline = rep.get("pipeline_time_ms", 0.0)
            dp_exposed = rep.get("dp_exposed_ms", 0.0)
            opt_time = rep.get("optimizer_time_ms", 0.0)
            opt_comm = rep.get("optimizer_comm_ms", 0.0)
            assert step == pytest.approx(pipeline + dp_exposed + opt_time + opt_comm, rel=1e-4), (
                f"dp={dp}: step_time({step:.4f}) != "
                f"pipeline({pipeline:.4f}) + dp_exposed({dp_exposed:.4f}) + "
                f"opt({opt_time:.4f}) + opt_comm({opt_comm:.4f})"
            )

    def test_pipeline_time_equals_warmup_steady_cooldown(self, dp_reports):
        """pipeline_time = warmup + steady + cooldown.

        Stack B (trace mode): pipeline_time does NOT include dp_exposed.
        dp_exposed is added to step_time separately, not inside pipeline_time.
        """
        for dp in (1, 4, 8):
            rep = dp_reports[dp]
            pipeline = rep.get("pipeline_time_ms", 0.0)
            warmup = rep.get("warmup_ms", 0.0)
            steady = rep.get("steady_ms", 0.0)
            cooldown = rep.get("cooldown_ms", 0.0)
            reconstructed = warmup + steady + cooldown
            assert pipeline == pytest.approx(reconstructed, rel=1e-4), (
                f"dp={dp}: pipeline_time({pipeline:.4f}) != "
                f"warmup+steady+cooldown({reconstructed:.4f})"
            )

    def test_dp_hidden_not_in_step_time(self, dp_reports):
        """dp_hidden is absorbed in bubble and does NOT increase step_time.

        If dp_hidden were fully exposed, step_time would be larger by dp_hidden.
        Verify step_time does NOT include dp_hidden as an extra additive term
        beyond the standard decomposition.
        """
        for dp in (4, 8):
            rep = dp_reports[dp]
            step = rep["step_time_ms"]
            pipeline = rep.get("pipeline_time_ms", 0.0)
            dp_exposed = rep.get("dp_exposed_ms", 0.0)
            dp_hidden = rep.get("dp_hidden_ms", 0.0)
            opt_time = rep.get("optimizer_time_ms", 0.0)
            opt_comm = rep.get("optimizer_comm_ms", 0.0)
            without_hidden = pipeline + dp_exposed + opt_time + opt_comm
            with_hidden = without_hidden + dp_hidden
            assert step == pytest.approx(without_hidden, rel=1e-4), (
                f"dp={dp}: step_time matches pipeline+dp_exposed+opt+opt_comm, "
                f"dp_hidden({dp_hidden:.4f}ms) is absorbed in bubble, not added"
            )
            if dp_hidden > 0:
                assert step < with_hidden, (
                    f"dp={dp}: step_time should be less than if dp_hidden were exposed"
                )

    def test_dp_hidden_positive_when_pp_gt1(self, dp_reports):
        """With PP=4 and dp_overlap_in_bubble=True, dp_hidden should be > 0.

        The pipeline bubble (warmup + cooldown) provides a window for DP
        communication to hide. With PP=4 there is significant bubble time
        and moderate comm volume that fits in the bubble.
        """
        for dp in (4, 8):
            rep = dp_reports[dp]
            dp_hidden = rep.get("dp_hidden_ms", 0.0)
            assert dp_hidden > 0, (
                f"dp={dp}: dp_hidden should be >0 with PP=4 bubble, got {dp_hidden}"
            )


# ── Excel: DP comm operator existence ──────────────────────────────────────

class TestDpCommExistence:
    """DP communication operator count and presence by dp value."""

    def test_count_matches_ddp_buckets(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        dp_rows = _filter_dp_comm(comm)
        bucket_ids = {
            str(row.get("Node ID", ""))
            for row in dp_rows
            if str(row.get("Node ID", "")).startswith("comm_grad_reduce_bucket_")
        }
        assert len(dp_rows) == len(bucket_ids) == 10, (
            f"Expected 10 DDP bucket comm ops, got rows={len(dp_rows)}, "
            f"bucket_ids={sorted(bucket_ids)}"
        )

    def test_none_for_dp1(self, dp_excel_data):
        comm = dp_excel_data[1].get("Communication Ops", [])
        dp_rows = _filter_dp_comm(comm)
        assert len(dp_rows) == 0, (
            f"Expected 0 DP comm ops for dp=1, got {len(dp_rows)}"
        )


# ── Excel: DP comm operator attributes ─────────────────────────────────────

class TestDpCommAttributes:
    """DP communication node attribute correctness."""

    def test_collective_is_reduce_scatter(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Collective Op") == "reduce_scatter", (
                f"Expected reduce_scatter, got {row.get('Collective Op')}"
            )

    def test_group_size_equals_dp(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Group Size") == 4, (
                f"Expected 4, got {row.get('Group Size')}"
            )

    def test_role_is_dp_grad_reduce(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Role") == "dp_grad_reduce", (
                f"Expected dp_grad_reduce, got {row.get('Role')}"
            )

    def test_scope_pattern(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            scope = str(row.get("Scope", ""))
            assert "data_parallel.grad_reduce.bucket_" in scope, (
                f"Unexpected scope: {scope}"
            )

    def test_inserted_by(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            inserted = str(row.get("Inserted By", "")).lower()
            assert "data_parallel" in inserted, (
                f"Expected data_parallel_pass, got {inserted}"
            )

    def test_has_stream_info(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            assert row.get("Stream Type"), (
                f"Missing Stream Type on {row.get('Node ID')}"
            )
            assert row.get("Stream ID") is not None, (
                f"Missing Stream ID on {row.get('Node ID')}"
            )

    def test_node_id_pattern(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        for row in _filter_dp_comm(comm):
            nid = str(row.get("Node ID", ""))
            assert nid.startswith("comm_grad_reduce_bucket_"), (
                f"Unexpected Node ID: {nid}"
            )


# ── Excel: DP comm scope encodes layer info ────────────────────────────────

class TestDpCommScope:
    """DP comm node scope carries bucket info in DDP bucket mode."""

    def test_scopes_contain_layer_keys(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        scopes = {str(r.get("Scope", "")) for r in _filter_dp_comm(comm)}
        assert scopes
        assert all("data_parallel.grad_reduce.bucket_" in s for s in scopes)

    def test_one_comm_per_scope(self, dp_excel_data):
        comm = dp_excel_data[4].get("Communication Ops", [])
        scopes = [str(r.get("Scope", "")) for r in _filter_dp_comm(comm)]
        assert len(scopes) == len(set(scopes)), (
            f"Duplicate scopes: {scopes}"
        )


# ── Excel: grad scale ops (aten.div.Scalar) ────────────────────────────────

class TestDpScaleOps:
    """DP gradient averaging scale nodes (aten.div.Scalar)
    in the Operators (fwd+bwd) sheet, bwd phase."""

    def _scale_rows(self, data: dict) -> list[dict]:
        t_ops = data.get("Operators (fwd+bwd)", [])
        return [
            r for r in t_ops
            if str(r.get("Phase", "")) == "bwd"
            and r.get("Op Type") == "aten.div.Scalar"
            and "grad_scale" in str(r.get("Node ID", ""))
        ]

    def test_count_eq_buckets(self, dp_excel_data):
        assert len(self._scale_rows(dp_excel_data[4])) == 10, (
            "Expected one grad scale op per DDP bucket"
        )

    def test_none_for_dp1(self, dp_excel_data):
        assert len(self._scale_rows(dp_excel_data[1])) == 0, (
            "Expected 0 scale ops for dp=1"
        )

    def test_node_id_pattern(self, dp_excel_data):
        for row in self._scale_rows(dp_excel_data[4]):
            nid = str(row.get("Node ID", ""))
            assert nid.startswith("grad_scale_bucket_"), (
                f"Unexpected scale node ID: {nid}"
            )


# ── PP Trace: DP communication nodes ──────────────────────────────────────

def _load_trace_events(trace_path: Path) -> list[dict]:
    """Load Chrome Trace JSON and return only X-duration events."""
    with open(trace_path) as f:
        data = json.load(f)
    return [
        evt for evt in data.get("traceEvents", [])
        if evt.get("ph") == "X"
    ]


def _filter_dp_comm_events(events: list[dict]) -> list[dict]:
    """Return only DP gradient reduction (reduce_scatter / all_reduce) events."""
    return [
        evt for evt in events
        if evt.get("args", {}).get("op_type", "") in (
            "comm.reduce_scatter", "comm.all_reduce",
        )
    ]


@pytest.fixture(scope="session")
def dp_pp_trace(dp_reports):
    """Session fixture: load pp_per_stage.json and pp_combined.json trace events."""
    traces = {}
    for dp in (1, 4):
        out_dir = dp_reports[dp].out_dir
        per_stage_path = out_dir / "pp_trace" / "pp_per_stage.json"
        combined_path = out_dir / "pp_trace" / "pp_combined.json"
        stitched_path = out_dir / "pp_trace" / "pp_stitched.json"
        traces[dp] = {
            "per_stage": _load_trace_events(per_stage_path) if per_stage_path.exists() else [],
            "combined": _load_trace_events(combined_path) if combined_path.exists() else [],
            "stitched": _load_trace_events(stitched_path) if stitched_path.exists() else [],
        }
    return traces


class TestPpTraceDpCommExistence:
    """DP communication nodes in PP trace output.

    Trace events are replicated per microbatch (M = global_batch / (micro_batch * dp)).
    For dp=4: M = 64 / (8*4) = 2, so 4 layers × 2 microbatches = 8 events.
    For dp=1: M = 64 / (8*1) = 8, but dp=1 has no DP comm, so 0 events.
    """

    def test_per_stage_count_eq_buckets_times_M(self, dp_pp_trace):
        """dp=4 with 4 layers and M=2 should produce 8 reduce_scatter events in pp_per_stage."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["per_stage"])
        M = 2
        expected = 10 * M
        assert len(dp_comm) == expected, (
            f"Expected {expected} DP comm events (4 layers × {M} microbatches) in pp_per_stage for dp=4, got {len(dp_comm)}"
        )

    def test_combined_count_eq_buckets_times_M(self, dp_pp_trace):
        """dp=4 with 4 layers and M=2 should produce 8 reduce_scatter events in pp_combined."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["combined"])
        M = 2
        expected = 10 * M
        assert len(dp_comm) == expected, (
            f"Expected {expected} DP comm events (4 layers × {M} microbatches) in pp_combined for dp=4, got {len(dp_comm)}"
        )

    def test_none_for_dp1_per_stage(self, dp_pp_trace):
        """dp=1 should have no reduce_scatter events in pp_per_stage."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[1]["per_stage"])
        assert len(dp_comm) == 0, (
            f"Expected 0 DP comm events for dp=1, got {len(dp_comm)}"
        )

    def test_none_for_dp1_combined(self, dp_pp_trace):
        """dp=1 should have no reduce_scatter events in pp_combined."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[1]["combined"])
        assert len(dp_comm) == 0, (
            f"Expected 0 DP comm events for dp=1, got {len(dp_comm)}"
        )

    def test_stitched_count_eq_buckets_times_M(self, dp_pp_trace):
        """pp_stitched also exposes detailed DP comm events."""
        assert len(_filter_dp_comm_events(dp_pp_trace[1]["stitched"])) == 0
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["stitched"])
        M = 2
        expected = 10 * M
        assert len(dp_comm) == expected, (
            f"Expected {expected} DP comm events in pp_stitched for dp=4, got {len(dp_comm)}"
        )


class TestPpTraceDpCommAttributes:
    """DP communication node attribute correctness in PP trace."""

    def test_phase_is_backward(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["per_stage"]):
            assert evt["args"]["phase"] == "bwd", (
                f"DP comm node should be in backward phase, got {evt['args']['phase']}"
            )

    def test_category_is_communication(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["per_stage"]):
            assert evt["cat"] == "communication", (
                f"DP comm node category should be 'communication', got {evt['cat']}"
            )

    def test_stream_type_is_comm(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["per_stage"]):
            assert evt["args"].get("stream_type") == "comm", (
                f"DP comm node stream_type should be 'comm', got {evt['args'].get('stream_type')}"
            )

    def test_duration_positive(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["per_stage"]):
            assert evt["dur"] > 0, (
                f"DP comm node duration should be > 0, got {evt['dur']}"
            )

    def test_name_contains_reduce_scatter(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["per_stage"]):
            assert "comm.reduce_scatter" in evt["name"], (
                f"DP comm node name should contain 'comm.reduce_scatter', got {evt['name']}"
            )

    def test_combined_has_view_detail(self, dp_pp_trace):
        for evt in _filter_dp_comm_events(dp_pp_trace[4]["combined"]):
            assert evt["args"].get("view") == "detail", (
                f"pp_combined DP comm node should have view='detail', got {evt['args'].get('view')}"
            )


class TestPpTraceDpCommStageDistribution:
    """DP comm nodes are distributed across PP stages based on layer assignment."""

    def test_dp_comm_spans_multiple_stages(self, dp_pp_trace):
        """With PP=4 and 4 layers, DP comm nodes should appear on at least 2 stages."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["per_stage"])
        pids = {evt["pid"] for evt in dp_comm}
        assert len(pids) >= 2, (
            f"DP comm nodes should span at least 2 PP stages, found stages: {pids}"
        )

    def test_dp_comm_count_per_stage_matches_buckets_times_M(self, dp_pp_trace):
        """Total DP comm nodes across all stages should equal layers × M."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["per_stage"])
        M = 2
        expected = 10 * M
        assert len(dp_comm) == expected, (
            f"Total DP comm nodes should equal {expected} (4 layers × {M} microbatches), got {len(dp_comm)}"
        )


class TestPpTraceDpCommTiming:
    """DP comm nodes should appear during the backward phase and overlap with compute."""

    def test_dp_comm_in_backward_timeline(self, dp_pp_trace):
        """All DP comm nodes should have timestamps > 0 (placed in the timeline)."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["per_stage"])
        for evt in dp_comm:
            assert evt["ts"] > 0, (
                f"DP comm node should have positive timestamp, got {evt['ts']}"
            )

    def test_dp_comm_duration_consistent_per_bucket(self, dp_pp_trace):
        """The same DDP bucket should have consistent duration across microbatches."""
        dp_comm = _filter_dp_comm_events(dp_pp_trace[4]["per_stage"])
        by_bucket: dict[int, list[float]] = {}
        for evt in dp_comm:
            bucket_index = evt.get("args", {}).get("bucket_index")
            assert bucket_index is not None, f"Missing bucket_index in event {evt}"
            by_bucket.setdefault(int(bucket_index), []).append(evt["dur"])

        assert len(by_bucket) == 10
        for bucket_index, durations in by_bucket.items():
            ref = durations[0]
            for d in durations[1:]:
                assert d == pytest.approx(ref, rel=0.1), (
                    f"Bucket {bucket_index} durations should be consistent: {durations}"
                )
