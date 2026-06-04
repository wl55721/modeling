"""Tests for ``_build_layer_display_map`` and ``_write_fwd_bwd_ops_sheet``.

Covers the two new interfaces added to ``python.zrt.transform.exporter``:

1. ``_build_layer_display_map(graph)`` — builds a layer-index → display-label
   map from ``LayerProfile`` metadata, including uniform-stride detection.
2. ``TrainingGraphExcelWriter._write_fwd_bwd_ops_sheet(...)`` — writes the
   merged ``Operators (fwd+bwd)`` sheet with a ``Phase`` column.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from python.zrt.graph.layer_strategy import LayerProfile, LayerType
from python.zrt.ir.graph import Edge, OpGraph
from python.zrt.ir.node import OpNode
from python.zrt.ir.types import DType, TensorMeta
from python.zrt.transform.exporter import (
    TrainingGraphExcelWriter,
    _build_layer_display_map,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _t(name: str, shape=(1,)) -> TensorMeta:
    return TensorMeta.from_shape_dtype(name, shape, DType.BF16)


def _mk_node(nid: str, layer: str = "", op_type: str = "aten.add.Tensor",
             scope: str = "", phase: str = "fwd") -> OpNode:
    n = OpNode(
        id=nid,
        op_type=op_type,
        inputs=[_t(f"{nid}_in")],
        outputs=[_t(f"{nid}_out")],
        scope=scope,
        layer=layer,
    )
    n.annotations["phase"] = phase
    return n


def _chain_graph(name: str, phase: str,
                 nodes: list[OpNode]) -> OpGraph:
    g = OpGraph(name=name, phase=phase)
    for n in nodes:
        g.nodes[n.id] = n
    for a, b in zip(nodes, nodes[1:]):
        g.edges.append(Edge(src=a.id, src_idx=0, dst=b.id, dst_idx=0,
                            tensor=a.outputs[0]))
    g._rebuild_adjacency()
    return g


def _make_profile(layer_types: list[LayerType],
                  typical_indices: list[int],
                  **counts) -> LayerProfile:
    defaults = dict(
        num_dense=0, num_moe=0,
        num_hca_hash=0, num_hca_topk=0, num_hca=0,
        num_csa_hash=0, num_csa_topk=0, num_csa=0,
        num_swa_hash=0, num_swa_topk=0, num_swa=0,
    )
    defaults.update(counts)
    return LayerProfile(
        layer_types=layer_types,
        typical_indices=typical_indices,
        **defaults,
    )


def _graph_with_profile(profile: LayerProfile) -> OpGraph:
    g = OpGraph(name="t", phase="train")
    g.metadata["layer_profile"] = profile
    g.metadata["typical_indices"] = profile.typical_indices
    return g


def _ctx():
    import python.zrt.hardware.registry as hw_registry
    from python.zrt.transform import ParallelConfig, StreamConfig, TransformContext
    hw = hw_registry.load("nvidia_h100_sxm")
    return TransformContext(
        hw_spec=hw,
        parallel=ParallelConfig(tp=1),
        stream_config=StreamConfig(num_compute_streams=1, num_comm_streams=1),
    )


def _sheet_rows(path: Path, sheet_name: str,
                phase: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in values):
                continue
            rows.append(dict(zip(header, values)))
        if phase is not None:
            rows = [r for r in rows if str(r.get("Phase", "")) == phase]
        return rows
    finally:
        wb.close()


# ═════════════════════════════════════════════════════════════════════════════
# _build_layer_display_map
# ═════════════════════════════════════════════════════════════════════════════

class TestBuildLayerDisplayMap:

    def test_empty_graph_returns_empty(self):
        g = OpGraph(name="t", phase="train")
        assert _build_layer_display_map(g) == {}

    def test_no_layer_profile_returns_empty(self):
        g = OpGraph(name="t", phase="train")
        g.metadata["typical_indices"] = [0]
        assert _build_layer_display_map(g) == {}

    def test_no_typical_indices_returns_empty(self):
        g = OpGraph(name="t", phase="train")
        g.metadata["layer_profile"] = _make_profile(
            [LayerType.DENSE], [0], num_dense=1)
        assert _build_layer_display_map(g) == {}

    def test_empty_typical_indices_returns_empty(self):
        g = OpGraph(name="t", phase="train")
        g.metadata["layer_profile"] = _make_profile(
            [LayerType.DENSE], [], num_dense=1)
        g.metadata["typical_indices"] = []
        assert _build_layer_display_map(g) == {}

    def test_single_layer_no_range(self):
        profile = _make_profile(
            [LayerType.DENSE, LayerType.MOE, LayerType.MOE],
            [0, 1],
            num_dense=1, num_moe=2,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0, 1层)"
        assert m["1"] == "1 (1-2, 2层)"
        assert m["2"] == "1 (1-2, 2层)"

    def test_consecutive_layers_no_step(self):
        profile = _make_profile(
            [LayerType.DENSE] * 5,
            [0],
            num_dense=5,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        for i in range(5):
            assert m[str(i)] == "0 (0-4, 5层)"

    def test_two_layers_no_step(self):
        """count=2 < 3 → step never shown."""
        profile = _make_profile(
            [LayerType.MOE, LayerType.MOE],
            [0],
            num_moe=2,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0-1, 2层)"
        assert m["1"] == "0 (0-1, 2层)"
        assert "step" not in m["0"]

    def test_uniform_step_2(self):
        """Alternating DENSE/MOE → step 2 for each type."""
        types = [LayerType.DENSE, LayerType.MOE] * 15  # 30 layers
        profile = _make_profile(types, [0, 1], num_dense=15, num_moe=15)
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0-28, step 2, 15层)"
        assert m["28"] == "0 (0-28, step 2, 15层)"
        assert m["1"] == "1 (1-29, step 2, 15层)"
        assert m["29"] == "1 (1-29, step 2, 15层)"

    def test_uniform_step_3(self):
        types = [LayerType.DENSE, LayerType.MOE, LayerType.HCA_HASH] * 10
        profile = _make_profile(
            types, [0, 1, 2],
            num_dense=10, num_moe=10, num_hca_hash=10,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0-27, step 3, 10层)"
        assert m["1"] == "1 (1-28, step 3, 10层)"
        assert m["2"] == "2 (2-29, step 3, 10层)"

    def test_non_uniform_diffs_no_step(self):
        """Indices [0, 2, 3] → diffs [2,1] → not uniform → no step."""
        types = [LayerType.DENSE, LayerType.MOE, LayerType.DENSE,
                 LayerType.DENSE, LayerType.MOE]
        profile = _make_profile(types, [0], num_dense=3, num_moe=2)
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0-3, 3层)"
        assert "step" not in m["0"]

    def test_step_1_consecutive_no_step_label(self):
        """step=1 (consecutive) should NOT show 'step 1'."""
        profile = _make_profile(
            [LayerType.DENSE] * 10,
            [0],
            num_dense=10,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert m["0"] == "0 (0-9, 10层)"
        assert "step" not in m["0"]

    def test_dsv4_like_profile(self):
        """Realistic DeepSeek-V4 profile: hca_hash(2) + csa_hash(1) +
        hca_topk(29, step 2) + csa_topk(29, step 2)."""
        types = (
            [LayerType.HCA_HASH, LayerType.HCA_HASH]
            + [LayerType.CSA_HASH]
            + [LayerType.HCA_TOPK, LayerType.CSA_TOPK] * 29
        )
        profile = _make_profile(
            types, [0, 2, 3, 4],
            num_hca_hash=2, num_csa_hash=1,
            num_hca_topk=29, num_csa_topk=29,
            num_hca=31, num_csa=30,
        )
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)

        assert m["0"] == "0 (0-1, 2层)"
        assert m["1"] == "0 (0-1, 2层)"
        assert m["2"] == "2 (2, 1层)"
        assert m["3"] == "3 (3-59, step 2, 29层)"
        assert m["59"] == "3 (3-59, step 2, 29层)"
        assert m["4"] == "4 (4-60, step 2, 29层)"
        assert m["60"] == "4 (4-60, step 2, 29层)"

    def test_all_layers_mapped(self):
        """Every layer index in the profile should appear in the map."""
        types = [LayerType.DENSE] * 4 + [LayerType.MOE] * 8
        profile = _make_profile(types, [0, 4], num_dense=4, num_moe=8)
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert set(m.keys()) == {str(i) for i in range(12)}

    def test_layer_without_typical_index_excluded(self):
        """A layer type not matching any typical index is left unmapped."""
        types = [LayerType.DENSE, LayerType.MOE, LayerType.HCA_HASH]
        profile = _make_profile(types, [0], num_dense=1, num_moe=1,
                                num_hca_hash=1)
        g = _graph_with_profile(profile)
        m = _build_layer_display_map(g)
        assert "0" in m
        assert "1" not in m
        assert "2" not in m


# ═════════════════════════════════════════════════════════════════════════════
# _write_fwd_bwd_ops_sheet
# ═════════════════════════════════════════════════════════════════════════════

class TestWriteFwdBwdOpsSheet:

    def _write_and_read(self, tmp_path: Path,
                        fwd_nodes: list[OpNode],
                        bwd_nodes: list[OpNode] | None = None,
                        layer_profile: LayerProfile | None = None,
                        ) -> tuple[list[dict], list[dict]]:
        fwd_g = _chain_graph("fwd", "train_forward", fwd_nodes)
        if layer_profile:
            fwd_g.metadata["layer_profile"] = layer_profile
            fwd_g.metadata["typical_indices"] = layer_profile.typical_indices

        bwd_g = None
        if bwd_nodes is not None:
            bwd_g = _chain_graph("bwd", "train_backward", bwd_nodes)
            if layer_profile:
                bwd_g.metadata["layer_profile"] = layer_profile
                bwd_g.metadata["typical_indices"] = layer_profile.typical_indices

        path = tmp_path / "test.xlsx"
        writer = TrainingGraphExcelWriter()
        wb = openpyxl.Workbook()
        writer._write_metadata_sheet(wb, fwd_g, _ctx())
        writer._write_fwd_bwd_ops_sheet(wb, fwd_g, bwd_g, _ctx())
        wb.save(path)

        fwd_rows = _sheet_rows(path, "Operators (fwd+bwd)", phase="fwd")
        bwd_rows = _sheet_rows(path, "Operators (fwd+bwd)", phase="bwd")
        return fwd_rows, bwd_rows

    def test_sheet_name(self, tmp_path):
        fwd = [_mk_node("a", layer="0", phase="fwd")]
        fwd_rows, _ = self._write_and_read(tmp_path, fwd)
        assert len(fwd_rows) == 1

    def test_phase_column(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        bwd = [_mk_node("b0", layer="0", phase="bwd")]
        fwd_rows, bwd_rows = self._write_and_read(tmp_path, fwd, bwd)
        assert all(r["Phase"] == "fwd" for r in fwd_rows)
        assert all(r["Phase"] == "bwd" for r in bwd_rows)

    def test_fwd_before_bwd(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd"),
               _mk_node("f1", layer="1", phase="fwd")]
        bwd = [_mk_node("b0", layer="0", phase="bwd"),
               _mk_node("b1", layer="1", phase="bwd")]
        path = tmp_path / "test.xlsx"
        fwd_g = _chain_graph("fwd", "train_forward", fwd)
        bwd_g = _chain_graph("bwd", "train_backward", bwd)
        writer = TrainingGraphExcelWriter()
        wb = openpyxl.Workbook()
        writer._write_metadata_sheet(wb, fwd_g, _ctx())
        writer._write_fwd_bwd_ops_sheet(wb, fwd_g, bwd_g, _ctx())
        wb.save(path)

        all_rows = _sheet_rows(path, "Operators (fwd+bwd)")
        phases = [r["Phase"] for r in all_rows]
        assert phases == ["fwd", "fwd", "bwd", "bwd"]

    def test_bwd_graph_none_produces_fwd_only(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        fwd_rows, bwd_rows = self._write_and_read(tmp_path, fwd, bwd_nodes=None)
        assert len(fwd_rows) == 1
        assert len(bwd_rows) == 0

    def test_layer_display_map_applied(self, tmp_path):
        """Layer column should show typical-layer label when profile exists."""
        types = [LayerType.DENSE, LayerType.MOE, LayerType.MOE, LayerType.MOE]
        profile = _make_profile(types, [0, 1], num_dense=1, num_moe=3)

        fwd = [
            _mk_node("f0", layer="0", phase="fwd"),
            _mk_node("f1", layer="1", phase="fwd"),
            _mk_node("f2", layer="2", phase="fwd"),
            _mk_node("f3", layer="3", phase="fwd"),
        ]
        fwd_rows, _ = self._write_and_read(tmp_path, fwd,
                                           layer_profile=profile)
        layers = [r["Layer"] for r in fwd_rows]
        assert layers[0] == "0 (0, 1层)"
        assert layers[1] == "1 (1-3, 3层)"
        assert layers[2] == "1 (1-3, 3层)"
        assert layers[3] == "1 (1-3, 3层)"

    def test_layer_display_step(self, tmp_path):
        """Step should appear when stride is uniform and > 1."""
        types = [LayerType.DENSE, LayerType.MOE] * 5  # 10 layers, step=2
        profile = _make_profile(types, [0, 1], num_dense=5, num_moe=5)

        fwd = [_mk_node(f"f{i}", layer=str(i), phase="fwd") for i in range(10)]
        fwd_rows, _ = self._write_and_read(tmp_path, fwd,
                                           layer_profile=profile)
        layers = [r["Layer"] for r in fwd_rows]
        assert layers[0] == "0 (0-8, step 2, 5层)"
        assert layers[1] == "1 (1-9, step 2, 5层)"

    def test_header_columns(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        path = tmp_path / "test.xlsx"
        fwd_g = _chain_graph("fwd", "train_forward", fwd)
        writer = TrainingGraphExcelWriter()
        wb = openpyxl.Workbook()
        writer._write_metadata_sheet(wb, fwd_g, _ctx())
        writer._write_fwd_bwd_ops_sheet(wb, fwd_g, None, _ctx())
        wb.save(path)

        wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb2["Operators (fwd+bwd)"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb2.close()

        assert header[0] == "Phase"
        assert "Node ID" in header
        assert "Op Type" in header
        assert "Layer" in header
        assert "Recompute" in header
        assert "Activation (B)" in header
        assert "Activation Memory (µs)" in header
        assert "Recompute Replay (µs)" in header
        assert "Final Latency (µs)" in header

    def test_recompute_flag_in_bwd(self, tmp_path):
        bwd_node = _mk_node("b0", layer="0", phase="bwd")
        bwd_node.annotations["recompute"] = True
        bwd_node.annotations["recompute_latency_us"] = 5.0
        bwd_node.annotations["base_latency_us"] = 10.0
        bwd_node.annotations["latency_us"] = 15.0

        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        _, bwd_rows = self._write_and_read(tmp_path, fwd, [bwd_node])
        assert bwd_rows[0]["Recompute"] == "YES"

    def test_no_recompute_flag(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        bwd = [_mk_node("b0", layer="0", phase="bwd")]
        _, bwd_rows = self._write_and_read(tmp_path, fwd, bwd)
        assert bwd_rows[0]["Recompute"] in ("", None)

    def test_activation_columns_fwd(self, tmp_path):
        fwd_node = _mk_node("f0", layer="0", phase="fwd")
        fwd_node.annotations["saved_activation_bytes"] = 1024
        fwd_node.annotations["activation_memory_us"] = 2.5
        fwd_rows, _ = self._write_and_read(tmp_path, [fwd_node])
        assert fwd_rows[0]["Activation (B)"] == 1024
        assert fwd_rows[0]["Activation Memory (µs)"] == 2.5

    def test_activation_columns_bwd_empty(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        bwd = [_mk_node("b0", layer="0", phase="bwd")]
        _, bwd_rows = self._write_and_read(tmp_path, fwd, bwd)
        assert bwd_rows[0]["Activation (B)"] in ("", None)
        assert bwd_rows[0]["Activation Memory (µs)"] in ("", None)

    def test_node_count_matches(self, tmp_path):
        fwd = [_mk_node(f"f{i}", layer=str(i), phase="fwd") for i in range(5)]
        bwd = [_mk_node(f"b{i}", layer=str(i), phase="bwd") for i in range(3)]
        fwd_rows, bwd_rows = self._write_and_read(tmp_path, fwd, bwd)
        assert len(fwd_rows) == 5
        assert len(bwd_rows) == 3

    def test_op_type_preserved(self, tmp_path):
        fwd = [
            _mk_node("f0", layer="0", phase="fwd", op_type="aten.mm.default"),
            _mk_node("f1", layer="0", phase="fwd", op_type="aten.add.Tensor"),
        ]
        fwd_rows, _ = self._write_and_read(tmp_path, fwd)
        ops = [r["Op Type"] for r in fwd_rows]
        assert "aten.mm.default" in ops
        assert "aten.add.Tensor" in ops

    def test_auto_filter_set(self, tmp_path):
        fwd = [_mk_node("f0", layer="0", phase="fwd")]
        path = tmp_path / "test.xlsx"
        fwd_g = _chain_graph("fwd", "train_forward", fwd)
        writer = TrainingGraphExcelWriter()
        wb = openpyxl.Workbook()
        writer._write_metadata_sheet(wb, fwd_g, _ctx())
        writer._write_fwd_bwd_ops_sheet(wb, fwd_g, None, _ctx())
        wb.save(path)

        wb2 = openpyxl.load_workbook(path)
        ws = wb2["Operators (fwd+bwd)"]
        assert ws.auto_filter.ref is not None
        assert ws.freeze_panes == "A2"
        wb2.close()
